from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.core.paginator import Paginator
from .models import Income, Category
from django.http import HttpResponse
from django.http import JsonResponse
import datetime


from io import BytesIO
from reportlab.pdfgen import canvas
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
import csv

@login_required(login_url='/authentication/login')
def index(request):
    categories = Category.objects.all()
    incomes = Income.objects.filter(owner=request.user)
    paginator = Paginator(incomes, 2)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    context = {
        'incomes': incomes,
        'page_obj': page_obj,
    }
    return render(request, 'income/index.html', context)

@login_required(login_url='/authentication/login')
def add_income(request):
    categories = Category.objects.all()
    context = {
        'categories': categories,
        'values': request.POST,
    }

    if request.method == "GET":
        return render(request, 'income/add-income.html', context)

    if request.method == "POST":
        amount=request.POST['amount']
        if not amount:
            messages.error(request,"Amount is required")
            return render(request,'income/add-income.html',context)
   
        description=request.POST['description']
        date=request.POST['income_date']
        category=request.POST['category']

        if not description:
            messages.error(request,"Description is required")
            return render(request,'income/add-income.html',context)

        Income.objects.create(owner=request.user,amount=amount,date=date,description=description,category=category)
        messages.success(request,"Income saved successfully")

        return redirect('income')
    

@login_required(login_url='/authentication/login')
def income_edit(request, id):
    income = Income.objects.get(pk=id)
    categories = Category.objects.all()
    context = {
        'income': income,
        'values': income,
        'categories': categories,
    }

    if request.method == 'GET':
        return render(request, 'income/edit-income.html', context)

    if request.method == "POST":
        amount=request.POST['amount']
        if not amount:
            messages.error(request,"Amount is required")
            return render(request,'income/edit-income.html',context)
   
        description=request.POST['description']
        date=request.POST['income_date']
        category=request.POST['category']

        if not description:
            messages.error(request,"Description is required")
            return render(request,'income/edit-income.html',context)

        income.owner=request.user
        income.amount=amount
        income.date=date
        income.category=category
        income.description=description

        income.save()
        messages.success(request,"Income updated successfully")

        return redirect('income')

@login_required(login_url='/authentication/login')
def delete_income(request, id):
    income = Income.objects.get(pk=id)
    income.delete()
    messages.success(request, "Income removed successfully")
    return redirect('income')

@login_required(login_url='/authentication/login')
def export_to_excel(request):
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename="incomes.xlsx"'

    incomes = Income.objects.filter(owner=request.user)

    workbook = Workbook()
    worksheet = workbook.active

    worksheet['A1'] = 'Date'
    worksheet['B1'] = 'Amount'
    worksheet['C1'] = 'Category'
    worksheet['D1'] = 'Description'

    for row_num, income in enumerate(incomes, start=2):
        worksheet.cell(row=row_num, column=1, value=income.date)
        worksheet.cell(row=row_num, column=2, value=income.amount)
        worksheet.cell(row=row_num, column=3, value=income.category)
        worksheet.cell(row=row_num, column=4, value=income.description)

    workbook.save(response)

    return response

@login_required(login_url='/authentication/login')
def export_to_pdf(request):
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="incomes.pdf"'

    incomes = Income.objects.filter(owner=request.user)

    buffer = BytesIO()
    p = canvas.Canvas(buffer)

    p.drawString(100, 750, 'Date')
    p.drawString(200, 750, 'Amount')
    p.drawString(300, 750, 'Category')
    p.drawString(400, 750, 'Description')

    y = 730
    for income in incomes:
        y -= 20
        p.drawString(100, y, str(income.date))
        p.drawString(200, y, str(income.amount))
        p.drawString(300, y, income.category)
        p.drawString(400, y, income.description)

    p.showPage()
    p.save()

    pdf = buffer.getvalue()
    buffer.close()
    response.write(pdf)

    return response

@login_required(login_url='/authentication/login')
def export_to_csv(request):
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="incomes.csv"'

    incomes = Income.objects.filter(owner=request.user)

    writer = csv.writer(response)
    writer.writerow(['Date', 'Amount', 'Category', 'Description'])

    for income in incomes:
        writer.writerow([income.date, income.amount, income.category, income.description])

    return response

def income_category_summary(request):
    today =datetime.date.today()
    six_months_ago = today - datetime.timedelta(days=30 * 6)
    incomes = Income.objects.filter(owner=request.user, date__gte=six_months_ago, date__lte=today)
    final_report = {}

    def get_category(income):
        return income.category 

    category_list = list(set(map(get_category, incomes)))

    def get_income_category_amount(category):
        amount = 0
        filtered_by_category = incomes.filter(category=category)

        for item in filtered_by_category:
            amount += item.amount

        return amount

    for x in incomes:
        for y in category_list:
            final_report[y] = get_income_category_amount(y)

    return JsonResponse({'income_category_data': final_report}, safe=False)

def stats_view2(request):
    return render(request,'income/income_summary.html')
