from django.shortcuts import render
from django.contrib.auth.decorators import login_required
from .models import Category,Expense
from django.contrib import messages
from django.shortcuts import redirect
from django.core.paginator import Paginator
import csv
from django.http import HttpResponse
from io import BytesIO
from reportlab.pdfgen import canvas
from django.http import HttpResponse
import openpyxl
from openpyxl.utils import get_column_letter
import json
from django.http import JsonResponse
import datetime


def search_expenses(request):
    if request.method == 'POST':
        search_str = json.loads(request.body).get('searchText')
        expenses = Expense.objects.filter(
            owner=request.user,
            amount__icontains=search_str,
        ) | Expense.objects.filter(
            owner=request.user,
            date__icontains=search_str,
        ) | Expense.objects.filter(
            owner=request.user,
            description__icontains=search_str,
        ) | Expense.objects.filter(
            owner=request.user,
            category__icontains=search_str,
        )
        data = list(expenses.values())  # Convert QuerySet to a list
        return JsonResponse(data, safe=False)

@login_required(login_url='/authentication/login')
def index(request):
    categories=Category.objects.all()
    expenses=Expense.objects.filter(owner=request.user)
    paginator=Paginator(expenses,2)
    page_number=request.GET.get('page')
    page_obj=Paginator.get_page(paginator,page_number)
    context={
        'expenses':expenses,
        'page_obj':page_obj
    }
    return render(request,'expenses/index.html',context) 

@login_required(login_url='/authentication/login')
def add_expense(request):
    categories=Category.objects.all()
    context={
        'categories':categories,
        'values':request.POST
    }  
    
    if request.method =="GET":
        return render(request,'expenses/add-expenses.html',context)

    if request.method=="POST":
        amount=request.POST['amount']

        if not amount:
            messages.error(request,"Amount is required")
            return render(request,'expenses/add-expenses.html',context)
   
        description=request.POST['description']
        date=request.POST['expense_date']
        category=request.POST['category']

        if not description:
            messages.error(request,"Description is required")
            return render(request,'expenses/add-expenses.html',context)

        Expense.objects.create(owner=request.user,amount=amount,date=date,description=description,category=category)
        messages.success(request,"Expense saved successfully")

        return redirect('expenses')
    
def expense_edit(request,id):
    expense=Expense.objects.get(pk=id)
    categories=Category.objects.all()
    context={
        'expense':expense,
        'values':expense,
        'categories':categories
    }
    
    if request.method =='GET':
        return render(request,'expenses/edit-expense.html',context)

    if request.method =='POST':
        amount=request.POST['amount']

        if not amount:
            messages.error(request,"Amount is required")
            return render(request,'expenses/edit-expense.html',context)
   
        description=request.POST['description']
        date=request.POST['expense_date']
        category=request.POST['category']

        if not description:
            messages.error(request,"Description is required")
            return render(request,'expenses/edit-expense.html',context)

        expense.owner=request.user
        expense.amount=amount
        expense.date=date
        expense.category=category
        expense.description=description

        expense.save()
        messages.success(request,"Expense updated successfully")

        return redirect('expenses')

def delete_expense(request,id):
    expense=Expense.objects.get(pk=id)
    expense.delete()
    messages.success(request,"Expense removed successfully")
    return redirect("expenses")

@login_required(login_url='/authentication/login')
def export_to_excel(request):
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename="expenses.xlsx"'

    expenses = Expense.objects.filter(owner=request.user)  # Adjust the query as needed

    workbook = openpyxl.Workbook()
    worksheet = workbook.active

    worksheet['A1'] = 'Date'
    worksheet['B1'] = 'Amount'
    worksheet['C1'] = 'Category'
    worksheet['D1'] = 'Description'

    for row_num, expense in enumerate(expenses, start=2):
        worksheet.cell(row=row_num, column=1, value=expense.date)
        worksheet.cell(row=row_num, column=2, value=expense.amount)
        worksheet.cell(row=row_num, column=3, value=expense.category)
        worksheet.cell(row=row_num, column=4, value=expense.description)

    workbook.save(response)

    return response

@login_required(login_url='/authentication/login')
def export_to_pdf(request):
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="expenses.pdf"'

    expenses = Expense.objects.filter(owner=request.user)  

    buffer = BytesIO()
    p = canvas.Canvas(buffer)

    p.drawString(100, 750, 'Date')
    p.drawString(200, 750, 'Amount')
    p.drawString(300, 750, 'Category')
    p.drawString(400, 750, 'Description')

    y = 730
    for expense in expenses:
        y -= 20
        p.drawString(100, y, str(expense.date))
        p.drawString(200, y, str(expense.amount))
        p.drawString(300, y, expense.category)
        p.drawString(400, y, expense.description)

    p.showPage()
    p.save()

    pdf = buffer.getvalue()
    buffer.close()
    response.write(pdf)

    return response


@login_required(login_url='/authentication/login')
def export_to_csv(request):
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="expenses.csv"'

    expenses = Expense.objects.filter(owner=request.user)  # Adjust the query as needed

    writer = csv.writer(response)
    writer.writerow(['Date', 'Amount', 'Category', 'Description'])

    for expense in expenses:
        writer.writerow([expense.date, expense.amount, expense.category, expense.description])

    return response


def expense_category_summary(request):
    today_date=datetime.date.today()
    six_months_ago=today_date-datetime.timedelta(days=30*6)
    expenses = Expense.objects.filter(owner=request.user, date__gte=six_months_ago, date__lte=today_date)
    finalrep={}

    def get_category(expense):
        return expense.category
    category_list=list(set(map(get_category,expenses)))


    def get_expense_category_amount(category):
        amount=0
        filtered_by_category=expenses.filter(category=category)

        for item in filtered_by_category:
            amount+=item.amount
        
        return amount
    
    for x in expenses:
        for y in category_list:
            finalrep[y]=get_expense_category_amount(y)

    return JsonResponse({'expense_category_data':finalrep},safe=False)

def stats_view(request):
    return render(request,'expenses/stats.html')

      



